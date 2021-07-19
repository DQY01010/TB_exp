import openpyxl


if __name__ == '__main__':
    wb_refine = openpyxl.load_workbook('./Tuberculosis_anno.xlsx')
    sheet_refine = wb_refine.worksheets[0]
    wb_label = openpyxl.Workbook()
    wb_label.create_sheet(index=0,title='sheet1')
    sheet_label = wb_label.worksheets[0]
    row_num = 0
    for i in range(sheet_refine.max_row+1):
#     for i in range(2): #test
        chars = []
        if row_num == 0:
            row_num += 1
        else:
            datapath = sheet_refine['H'+str(i)].value
            overall_id = sheet_refine['D'+str(i)].value
            if sheet_refine['I'+str(i)].value == None:
                density = ''
            else:
                density = sheet_refine['I'+str(i)].value
            if sheet_refine['L'+str(i)].value == None:
                char1 = ''
            else:
                char1 = str(sheet_refine['L'+str(i)].value)
            if sheet_refine['M'+str(i)].value == None:
                char2 = ''
            else:
                char2 = sheet_refine['M'+str(i)].value
            if sheet_refine['N'+str(i)].value == None:
                char3 = ''
            else:
                char3 = sheet_refine['N'+str(i)].value
            if sheet_refine['O'+str(i)].value == None:
                char4 = ''
            else:
                char4 = sheet_refine['O'+str(i)].value
            if sheet_refine['P'+str(i)].value == None:
                char5 = ''
            else:
                char5 = sheet_refine['P'+str(i)].value
            if sheet_refine['Q'+str(i)].value == None:
                char6  =''
            else:
                char6 = sheet_refine['Q'+str(i)].value
            if sheet_refine['R'+str(i)].value == None:
                char7 = ''
            else:
                char7 = sheet_refine['R'+str(i)].value
            if sheet_refine['S'+str(i)].value == None:
                char8 = ''
            else:
                char8 = sheet_refine['S'+str(i)].value
            if sheet_refine['T'+str(i)].value == None:
                char9 = ''
            else:
                char9 = sheet_refine['T'+str(i)].value
            if sheet_refine['U'+str(i)].value == None:
                char10 = ''
            else:
                char10 = sheet_refine['U'+str(i)].value
            if sheet_refine['V'+str(i)].value == None:
                char11 = ''
            else:
                char11 = sheet_refine['V'+str(i)].value
            if sheet_refine['W'+str(i)].value == None:
                char12 = ''
            else:
                char12 = sheet_refine['W'+str(i)].value
            chars.append(density)
            chars.append(char1)
            chars.append(char2)
            chars.append(char3)
            chars.append(char4)
            chars.append(char5)
            chars.append(char6)
            chars.append(char7)
            chars.append(char8)
            chars.append(char9)
            chars.append(char10)
            chars.append(char11)
            print(chars)
            np.save('/data2/duqy/DeepPhthisis/BenMalData/data/tb_210301_refine_label/'
                    +datapath.split('/')[-1]+'#'+overall_id.split('I')[0]+'_I'+overall_id.split('I')[1]+'.npy' ,chars)
            